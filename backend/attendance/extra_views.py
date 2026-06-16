"""
extra_views.py — Additional API Views
======================================
- StudentAttendanceView     : student portal (attendance per course)
- RegistrationWindowView    : admin opens/closes registration
- AttendanceExcelExportView : export .xlsx file
"""

from __future__ import annotations

import io
import logging
from datetime import date

from django.http import HttpResponse
from django.utils import timezone
from rest_framework.request import Request
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework import status

logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# 1.  Student Portal — attendance per course
# ---------------------------------------------------------------------------

class StudentAttendanceView(APIView):
    """
    GET /api/students/<student_id>/attendance/

    Returns attendance summary per enrolled course for a student.
    """

    def get(self, request: Request, student_id: str) -> Response:
        from .models import Attendance, AttendanceStatus, LectureSession, Student

        try:
            student = Student.objects.prefetch_related("courses").get(
                student_id=student_id
            )
        except Student.DoesNotExist:
            return Response(
                {"error": f"Student '{student_id}' not found."},
                status=status.HTTP_404_NOT_FOUND,
            )

        courses_data = []
        for course in student.courses.all():
            records        = Attendance.objects.filter(student=student, course=course)
            total_sessions = LectureSession.objects.filter(
                course=course, status="done"
            ).count()
            present = records.filter(status=AttendanceStatus.PRESENT).count()
            absent  = records.filter(status=AttendanceStatus.ABSENT).count()
            pct     = round(present / total_sessions * 100, 1) if total_sessions else 0

            last_rec = records.order_by("-date").first()

            courses_data.append({
                "course_id":      course.course_id,
                "course_name":    course.course_name,
                "doctor_name":    course.doctor_name,
                "present":        present,
                "absent":         absent,
                "total_sessions": total_sessions,
                "attendance_pct": pct,
                "at_risk":        pct < 75 and total_sessions > 0,
                "last_status":    last_rec.status if last_rec else "N/A",
            })

        return Response({
            "student_id":   student.student_id,
            "student_name": student.name,
            "department":   student.get_department_display(),
            "courses":      courses_data,
        })


# ---------------------------------------------------------------------------
# 2.  Registration Window (Admin)
# ---------------------------------------------------------------------------

class RegistrationWindowView(APIView):
    """
    GET  /api/admin/registration-window/  → returns { open: bool, closed_at }
    POST /api/admin/registration-window/  → body: { open: bool }
    """

    WINDOW_KEY = "registration_window_open"

    def _get_state(self):
        """Use a simple DB-backed config via Django cache or a JSON file."""
        import json
        from pathlib import Path
        cfg_path = Path(__file__).parent / "reg_window.json"
        if cfg_path.exists():
            return json.loads(cfg_path.read_text())
        return {"open": True, "closed_at": None}

    def _set_state(self, open_: bool):
        import json
        from pathlib import Path
        cfg_path = Path(__file__).parent / "reg_window.json"
        state = {
            "open":      open_,
            "closed_at": None if open_ else timezone.now().isoformat(),
        }
        cfg_path.write_text(json.dumps(state))
        return state

    def get(self, request: Request) -> Response:
        return Response(self._get_state())

    def post(self, request: Request) -> Response:
        open_ = request.data.get("open")
        if open_ is None:
            return Response({"error": "'open' (bool) is required."},
                            status=status.HTTP_400_BAD_REQUEST)
        state = self._set_state(bool(open_))
        action = "opened" if open_ else "closed"
        logger.info("Registration window %s by admin", action)
        return Response({
            "message": f"Registration window {action} successfully.",
            **state,
        })


def is_registration_open() -> bool:
    """Helper used in StudentRegistrationView to check window state."""
    import json
    from pathlib import Path
    cfg_path = Path(__file__).parent / "reg_window.json"
    if cfg_path.exists():
        return json.loads(cfg_path.read_text()).get("open", True)
    return True


# ---------------------------------------------------------------------------
# 3.  Excel Export — openpyxl
# ---------------------------------------------------------------------------

class AttendanceExcelExportView(APIView):
    """
    GET /api/attendance/export/?course_id=CS401

    Returns an .xlsx file with full attendance report for the course.
    """

    def get(self, request: Request) -> HttpResponse:
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError:
            return Response(
                {"error": "openpyxl not installed. Run: pip install openpyxl"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

        from .models import Attendance, AttendanceStatus, Course, LectureSession

        course_id = request.query_params.get("course_id", "").strip()
        if not course_id:
            return Response({"error": "'course_id' is required."},
                            status=status.HTTP_400_BAD_REQUEST)

        try:
            course = Course.objects.get(course_id=course_id)
        except Course.DoesNotExist:
            return Response({"error": f"Course '{course_id}' not found."},
                            status=status.HTTP_404_NOT_FOUND)

        # ── Build data ─────────────────────────────────────────────────────
        students       = list(course.enrolled_students.order_by("name"))
        total_sessions = LectureSession.objects.filter(
            course=course, status="done"
        ).count()

        # ── Create workbook ────────────────────────────────────────────────
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"{course_id} Attendance"

        # Styles
        HEADER_FILL  = PatternFill("solid", fgColor="1E3A5F")
        PRESENT_FILL = PatternFill("solid", fgColor="DCFCE7")
        ABSENT_FILL  = PatternFill("solid", fgColor="FEE2E2")
        RISK_FILL    = PatternFill("solid", fgColor="FEF3C7")
        BOLD         = Font(bold=True)
        WHITE_BOLD   = Font(bold=True, color="FFFFFF")
        CENTER       = Alignment(horizontal="center", vertical="center")
        thin         = Side(style="thin", color="D1D5DB")
        BORDER       = Border(left=thin, right=thin, top=thin, bottom=thin)

        # ── Title row ──────────────────────────────────────────────────────
        ws.merge_cells("A1:G1")
        ws["A1"] = f"Attendance Report — {course.course_name} ({course_id})"
        ws["A1"].font      = Font(bold=True, size=14, color="1E3A5F")
        ws["A1"].alignment = CENTER

        ws.merge_cells("A2:G2")
        ws["A2"] = f"Dr. {course.doctor_name}  |  Total Sessions: {total_sessions}  |  Generated: {date.today()}"
        ws["A2"].alignment = CENTER
        ws["A2"].font      = Font(size=10, color="6B7280")

        ws.append([])   # blank row

        # ── Header row ─────────────────────────────────────────────────────
        headers = ["#", "Student ID", "Student Name", "Department",
                   "Present", "Absent", "Attendance %", "Status"]
        ws.append(headers)
        for cell in ws[4]:
            cell.fill      = HEADER_FILL
            cell.font      = WHITE_BOLD
            cell.alignment = CENTER
            cell.border    = BORDER

        # ── Data rows ──────────────────────────────────────────────────────
        for i, student in enumerate(students, 1):
            records = Attendance.objects.filter(student=student, course=course)
            present = records.filter(status=AttendanceStatus.PRESENT).count()
            absent  = records.filter(status=AttendanceStatus.ABSENT).count()
            pct     = round(present / total_sessions * 100, 1) if total_sessions else 0
            at_risk = pct < 75

            row = [
                i,
                student.student_id,
                student.name,
                student.get_department_display(),
                present,
                absent,
                f"{pct}%",
                "⚠️ At Risk" if at_risk else "✅ Safe",
            ]
            ws.append(row)

            row_num = ws.max_row
            fill = RISK_FILL if at_risk else (
                PRESENT_FILL if pct >= 75 else ABSENT_FILL
            )
            for cell in ws[row_num]:
                cell.border    = BORDER
                cell.alignment = CENTER
                if cell.column in (5, 6, 7, 8):   # numeric/status cols
                    cell.fill = fill

        # ── Column widths ──────────────────────────────────────────────────
        widths = [5, 14, 28, 22, 10, 10, 14, 14]
        for col, width in zip(ws.columns, widths):
            ws.column_dimensions[col[0].column_letter].width = width

        # ── Freeze header ──────────────────────────────────────────────────
        ws.freeze_panes = "A5"

        # ── Serve file ─────────────────────────────────────────────────────
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        filename = f"attendance_{course_id}_{date.today()}.xlsx"
        response = HttpResponse(
            buffer.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response
