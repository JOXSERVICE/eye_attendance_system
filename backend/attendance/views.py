"""
views.py — University Smart Attendance System (Full Version)
"""
from __future__ import annotations
import logging, os, tempfile
from datetime import date
from io import BytesIO
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from django.db import transaction
from django.http import HttpResponse
from django.utils import timezone
from rest_framework import status
from rest_framework.parsers import FormParser, MultiPartParser
from rest_framework.request import Request
from rest_framework.response import Response
from rest_framework.views import APIView

from .models import Attendance, AttendanceStatus, Course, LectureSession, Student, RegistrationWindow
from .tasks import enroll_student_face, process_lecture_session

logger = logging.getLogger(__name__)
ALLOWED_IMAGE_TYPES = {".jpg", ".jpeg", ".png", ".bmp", ".webp"}
MAX_IMAGE_MB = 20

def _save_temp(image_file) -> Path:
    suffix = Path(image_file.name).suffix.lower()
    if suffix not in ALLOWED_IMAGE_TYPES:
        raise ValueError(f"Unsupported format '{suffix}'.")
    if image_file.size > MAX_IMAGE_MB * 1024 * 1024:
        raise ValueError(f"Image exceeds {MAX_IMAGE_MB} MB limit.")
    fd, path = tempfile.mkstemp(suffix=suffix, prefix="face_att_")
    with os.fdopen(fd, "wb") as f:
        for chunk in image_file.chunks():
            f.write(chunk)
    return Path(path)


# 1. Registration Window Status
class RegistrationWindowStatusView(APIView):
    """GET /api/registration/status/"""
    def get(self, request):
        window = RegistrationWindow.objects.first()
        if not window:
            return Response({"is_open": False, "message": "Registration window not configured."})
        return Response({
            "is_open":   window.is_open,
            "opened_at": window.opened_at,
            "closed_at": window.closed_at,
            "message":   "Registration is open." if window.is_open else "Registration window is closed.",
        })


# 2. Student Self-Registration
class StudentRegistrationView(APIView):
    """POST /api/students/register/"""
    parser_classes = [MultiPartParser, FormParser]

    def post(self, request):
        window = RegistrationWindow.objects.first()
        if not window or not window.is_open:
            return Response(
                {"error": "Registration is currently closed. Please contact the admin."},
                status=status.HTTP_403_FORBIDDEN,
            )
        for field in ["student_id", "name", "email", "department"]:
            if not request.data.get(field):
                return Response({"error": f"'{field}' is required."}, status=400)

        photo = request.FILES.get("photo")
        if not photo:
            return Response({"error": "A clear photo is required."}, status=400)

        student_id = request.data["student_id"].strip()
        if Student.objects.filter(student_id=student_id).exists():
            return Response({"error": f"Student '{student_id}' already registered."}, status=400)

        try:
            tmp_path = _save_temp(photo)
        except ValueError as e:
            return Response({"error": str(e)}, status=400)

        try:
            with transaction.atomic():
                student = Student.objects.create(
                    student_id = student_id,
                    name       = request.data["name"].strip(),
                    email      = request.data["email"].strip(),
                    department = request.data.get("department", "CS"),
                    is_niqabi  = request.data.get("is_niqabi", "false").lower() == "true",
                    photo      = photo,
                )
            enroll_student_face.delay(student.student_id, str(tmp_path))
            return Response({
                "message":    f"Student '{student.name}' registered successfully.",
                "student_id": student.student_id,
                "note":       "Face embedding is being extracted in the background.",
            }, status=201)
        except Exception as e:
            tmp_path.unlink(missing_ok=True)
            logger.exception("Registration failed: %s", e)
            return Response({"error": "Registration failed."}, status=500)


# 3. Professor Uploads Lecture Photo
class LectureSessionUploadView(APIView):
    """POST /api/sessions/upload/"""
    parser_classes = [MultiPartParser, FormParser]

    def post(self, request):
        course_id     = request.data.get("course_id", "").strip()
        uploaded_by   = request.data.get("uploaded_by", "").strip()
        lecture_image = request.FILES.get("lecture_image")

        if not course_id:   return Response({"error": "'course_id' is required."}, status=400)
        if not uploaded_by: return Response({"error": "'uploaded_by' is required."}, status=400)
        if not lecture_image: return Response({"error": "Lecture image is required."}, status=400)

        try:
            course = Course.objects.get(course_id=course_id)
        except Course.DoesNotExist:
            return Response({"error": f"Course '{course_id}' not found."}, status=404)

        date_str = request.data.get("date", "")
        try:
            session_date = date.fromisoformat(date_str) if date_str else date.today()
        except ValueError:
            return Response({"error": "Invalid date. Use YYYY-MM-DD."}, status=400)

        session = LectureSession.objects.create(
            course=course, date=session_date,
            lecture_image=lecture_image, uploaded_by=uploaded_by, status="pending",
        )
        process_lecture_session.delay(session.id)

        return Response({
            "message":    "Image uploaded. AI processing started.",
            "session_id": session.id,
            "course_id":  course_id,
            "date":       str(session_date),
            "status":     "pending",
        }, status=202)


# 4. Session Status
class SessionStatusView(APIView):
    """GET /api/sessions/<session_id>/status/"""
    def get(self, request, session_id):
        try:
            session = LectureSession.objects.get(pk=session_id)
        except LectureSession.DoesNotExist:
            return Response({"error": "Session not found."}, status=404)

        data = {
            "session_id":   session.id,
            "course_id":    session.course.course_id,
            "date":         str(session.date),
            "status":       session.status,
            "processed_at": session.processed_at,
        }
        if session.status == "done":
            data["present_count"] = session.attendances.filter(status=AttendanceStatus.PRESENT).count()
            data["absent_count"]  = session.attendances.filter(status=AttendanceStatus.ABSENT).count()
        return Response(data)


# 5. Daily Attendance List
class AttendanceListView(APIView):
    """GET /api/attendance/?course_id=CS401&date=2025-01-15"""
    def get(self, request):
        course_id = request.query_params.get("course_id", "").strip()
        date_str  = request.query_params.get("date", "").strip()
        if not course_id or not date_str:
            return Response({"error": "Both 'course_id' and 'date' are required."}, status=400)
        try:
            filter_date = date.fromisoformat(date_str)
        except ValueError:
            return Response({"error": "Invalid date format."}, status=400)

        records = (Attendance.objects.filter(course__course_id=course_id, date=filter_date)
                   .select_related("student").order_by("student__name"))
        data = [{
            "student_id":       r.student.student_id,
            "student_name":     r.student.name,
            "department":       r.student.get_department_display(),
            "status":           r.status,
            "time_in":          str(r.time_in),
            "similarity_score": r.similarity_score,
        } for r in records]

        return Response({
            "course_id":     course_id,
            "date":          date_str,
            "total":         len(data),
            "present_count": sum(1 for r in data if r["status"] == "Present"),
            "absent_count":  sum(1 for r in data if r["status"] == "Absent"),
            "records":       data,
        })


# 6. Full Report
class AttendanceReportView(APIView):
    """GET /api/attendance/report/?course_id=CS401"""
    def get(self, request):
        course_id = request.query_params.get("course_id", "").strip()
        if not course_id:
            return Response({"error": "'course_id' is required."}, status=400)
        try:
            course = Course.objects.get(course_id=course_id)
        except Course.DoesNotExist:
            return Response({"error": f"Course '{course_id}' not found."}, status=404)

        total_sessions = LectureSession.objects.filter(course=course, status="done").count()
        report = []
        for student in course.enrolled_students.all():
            records = Attendance.objects.filter(student=student, course=course)
            present = records.filter(status=AttendanceStatus.PRESENT).count()
            absent  = records.filter(status=AttendanceStatus.ABSENT).count()
            pct     = round(present / total_sessions * 100, 1) if total_sessions > 0 else 0
            report.append({
                "student_id":     student.student_id,
                "student_name":   student.name,
                "department":     student.get_department_display(),
                "present":        present,
                "absent":         absent,
                "total_sessions": total_sessions,
                "attendance_pct": pct,
                "at_risk":        pct < 75,
            })
        return Response({
            "course_id":   course_id,
            "course_name": course.course_name,
            "doctor_name": course.doctor_name,
            "report":      sorted(report, key=lambda x: x["attendance_pct"]),
        })


# 7. Export Excel
class AttendanceExportView(APIView):
    """GET /api/attendance/export/?course_id=CS401&date=2025-01-15&mode=daily|report"""
    def get(self, request):
        course_id = request.query_params.get("course_id", "").strip()
        date_str  = request.query_params.get("date", "").strip()
        mode      = request.query_params.get("mode", "daily")

        try:
            course = Course.objects.get(course_id=course_id)
        except Course.DoesNotExist:
            return HttpResponse("Course not found.", status=404)

        wb = openpyxl.Workbook()
        ws = wb.active

        header_font  = Font(bold=True, color="FFFFFF", size=12)
        header_fill  = PatternFill("solid", fgColor="1E3A5F")
        present_fill = PatternFill("solid", fgColor="DCFCE7")
        absent_fill  = PatternFill("solid", fgColor="FEE2E2")
        risk_fill    = PatternFill("solid", fgColor="FEF3C7")
        center       = Alignment(horizontal="center", vertical="center")

        if mode == "daily" and date_str:
            ws.title = f"Attendance {date_str}"
            try:
                filter_date = date.fromisoformat(date_str)
            except ValueError:
                return HttpResponse("Invalid date.", status=400)

            ws.merge_cells("A1:G1")
            ws["A1"] = f"Attendance Sheet — {course.course_name} — {date_str}"
            ws["A1"].font = Font(bold=True, size=14, color="1E3A5F")
            ws["A1"].alignment = center
            ws.merge_cells("A2:G2")
            ws["A2"] = f"Dr. {course.doctor_name}  |  Course: {course_id}"
            ws["A2"].alignment = center

            for col, h in enumerate(["#","Student ID","Name","Department","Time In","Similarity %","Status"], 1):
                cell = ws.cell(row=3, column=col, value=h)
                cell.font = header_font; cell.fill = header_fill; cell.alignment = center

            records = (Attendance.objects.filter(course=course, date=filter_date)
                       .select_related("student").order_by("student__name"))
            for i, r in enumerate(records, 1):
                row_fill = present_fill if r.status == "Present" else absent_fill
                for col, val in enumerate([
                    i, r.student.student_id, r.student.name,
                    r.student.get_department_display(), str(r.time_in),
                    f"{r.similarity_score*100:.1f}%" if r.similarity_score else "—",
                    r.status
                ], 1):
                    cell = ws.cell(row=i+3, column=col, value=val)
                    cell.fill = row_fill; cell.alignment = center

            for col, w in zip("ABCDEFG", [5,15,25,22,12,14,12]):
                ws.column_dimensions[col].width = w
            filename = f"attendance_{course_id}_{date_str}.xlsx"

        else:
            ws.title = "Attendance Report"
            total_sessions = LectureSession.objects.filter(course=course, status="done").count()
            ws.merge_cells("A1:H1")
            ws["A1"] = f"Attendance Report — {course.course_name}"
            ws["A1"].font = Font(bold=True, size=14, color="1E3A5F")
            ws["A1"].alignment = center
            ws.merge_cells("A2:H2")
            ws["A2"] = f"Dr. {course.doctor_name}  |  Total Sessions: {total_sessions}"
            ws["A2"].alignment = center

            for col, h in enumerate(["#","Student ID","Name","Department","Present","Absent","Attendance %","Status"], 1):
                cell = ws.cell(row=3, column=col, value=h)
                cell.font = header_font; cell.fill = header_fill; cell.alignment = center

            for i, student in enumerate(course.enrolled_students.all().order_by("name"), 1):
                records = Attendance.objects.filter(student=student, course=course)
                present = records.filter(status=AttendanceStatus.PRESENT).count()
                absent  = records.filter(status=AttendanceStatus.ABSENT).count()
                pct     = round(present/total_sessions*100, 1) if total_sessions > 0 else 0
                at_risk = pct < 75
                for col, val in enumerate([
                    i, student.student_id, student.name,
                    student.get_department_display(),
                    present, absent, f"{pct}%",
                    "⚠ At Risk" if at_risk else "✓ Safe"
                ], 1):
                    cell = ws.cell(row=i+3, column=col, value=val)
                    cell.fill = risk_fill if at_risk else present_fill
                    cell.alignment = center

            for col, w in zip("ABCDEFGH", [5,15,25,22,10,10,14,12]):
                ws.column_dimensions[col].width = w
            filename = f"report_{course_id}.xlsx"

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        response = HttpResponse(
            buffer.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response


# 8. Student Portal — Personal Summary
class StudentSummaryView(APIView):
    """GET /api/student/<student_id>/summary/"""
    def get(self, request, student_id):
        try:
            student = Student.objects.get(student_id=student_id)
        except Student.DoesNotExist:
            return Response({"error": "Student not found."}, status=404)

        courses_data = []
        for course in student.courses.all():
            total_sessions = LectureSession.objects.filter(course=course, status="done").count()
            records = Attendance.objects.filter(student=student, course=course)
            present = records.filter(status=AttendanceStatus.PRESENT).count()
            absent  = records.filter(status=AttendanceStatus.ABSENT).count()
            pct     = round(present/total_sessions*100, 1) if total_sessions > 0 else 0
            recent  = records.order_by("-date")[:5]
            courses_data.append({
                "course_id":      course.course_id,
                "course_name":    course.course_name,
                "doctor_name":    course.doctor_name,
                "present":        present,
                "absent":         absent,
                "total_sessions": total_sessions,
                "attendance_pct": pct,
                "at_risk":        pct < 75,
                "recent_records": [
                    {"date": str(r.date), "status": r.status, "time_in": str(r.time_in)}
                    for r in recent
                ],
            })

        return Response({
            "student_id":   student.student_id,
            "student_name": student.name,
            "department":   student.get_department_display(),
            "courses":      courses_data,
        })


# 9. Courses List
class CourseListView(APIView):
    """GET /api/courses/"""
    def get(self, request):
        return Response([{
            "course_id":     c.course_id,
            "course_name":   c.course_name,
            "doctor_name":   c.doctor_name,
            "student_count": c.student_count,
        } for c in Course.objects.all()])
